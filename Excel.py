from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date
import os
from Kod_2 import *

# функция 2: записывает в файл список участников по категории
def Spisok_excel(ct, filename="spisok_excel.txt"):
    try:
        with open(filename, "a", encoding="utf-8") as file:
            for uchastnik in ct.people:
                file.write(str(uchastnik) + str(ct.name) + "\n")  # Запись информации об участнике с новой строкой
    except Exception as fall:
        print(f"Ошибка при записи в файл: {fall}")

def students_to_excel(self):
    try:
        with open('spisok_excel.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()

        wb = Workbook()
        ws = wb.active
        today = date.today()
        ws['A1'] = f"Список участников на {today}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        headers = ["ID", "Фамилия", "Имя", "Отчество", "Дата рождения", "Пол", "Телефон", "Телеграм", "Статус участия", "Категория"]
        ws.append(headers)

        for line in lines:
            st = line.strip().split(' ')
            ws.append(st)
        wb.save('new_excel.xlsx')
        print(f"Данные экспортированы в файл new_excel.xlsx")

    except Exception as e:(
        print(f"Ошибка: {e}"))


students_to_excel(MR.categories)

